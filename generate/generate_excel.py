
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def generate_excel(summary):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Summary"

    ws.append(["Date", "Income", "Expense", "Net"])

    red_fill = PatternFill(start_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", fill_type="solid")

    for row in summary:
        ws.append([
            row["date"],
            row["income"],
            row["expense"],
            row["net"]
        ])

        last_row = ws.max_row
        net = row["net"]

        # ใส่สี
        if net < 0:
            for col in range(1, 5):
                ws.cell(row=last_row, column=col).fill = red_fill
        else:
            for col in range(1, 5):
                ws.cell(row=last_row, column=col).fill = green_fill

    file_path = "statement_summary.xlsx"
    wb.save(file_path)

    return file_path